import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from models.student import Student
from models.task import Task


class StepikProgressParser:

    def __init__(self, path_to_file: str):
        wb = openpyxl.load_workbook(path_to_file)
        self.worksheet: Worksheet = wb.active

    def parse(self) -> list[Student]:
        students = self.__parse_names()
        self.__parse_tasks(students)
        return students

    def __parse_names(self) -> list[Student]:
        students: list[Student] = []
        user_id: int | None = 0
        row = 2
        count = 0
        while user_id is not None:
            user_id = self.worksheet.cell(row=row, column=1).value
            last_name = self.worksheet.cell(row=row, column=2).value
            first_name = self.worksheet.cell(row=row, column=3).value

            if last_name is None:
                last_name = f'Uncheck_{count}'
                count += 1

            if first_name is None:
                first_name = ''

            if user_id is not None:
                students.append(Student(user_id, last_name, first_name))

            row += 1
        return students

    def __parse_tasks(self, students: list[Student]) -> None:
        col = 4
        finish_col_name = 'total'
        col_name = ''
        task_name = ''
        task_step = ''
        while col_name != finish_col_name:
            row = 1
            data = ''
            while data is not None:
                data = self.worksheet.cell(row=row, column=col).value
                if data is None:
                    break

                if data == finish_col_name:
                    col_name = finish_col_name
                    break

                if row <= 1:
                    task = data.split('Q')
                    task_name = task[0].rstrip()
                    task_step = 'Q' + task[1]
                else:
                    if task_name not in students[row - 2].tasks:
                        students[row - 2].tasks[task_name] = []
                    students[row - 2].tasks[task_name].append(
                        Task(task_step, data > 0)
                    )
                row += 1
            col += 1





